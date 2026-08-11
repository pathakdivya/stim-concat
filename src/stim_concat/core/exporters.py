"""Writing timelines, settings and build summaries to disk."""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable, Sequence
from datetime import datetime, timezone
from pathlib import Path

from .config import BuildConfig
from .timeline import TIMELINE_COLUMNS, Timeline

__all__ = [
    "XlsxUnavailable",
    "write_settings_json",
    "write_summary_csv",
    "write_summary_markdown",
    "write_timeline_csv",
    "write_timeline_xlsx",
]


class XlsxUnavailable(RuntimeError):
    """Raised when openpyxl is not installed."""


def write_timeline_csv(timeline: Timeline, path: str | Path) -> Path:
    """Write the annotated event timeline as CSV."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(TIMELINE_COLUMNS))
        writer.writeheader()
        writer.writerows(timeline.rows())
    return path


def write_timeline_xlsx(timeline: Timeline, path: str | Path) -> Path:
    """Write the timeline as a formatted single-sheet workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise XlsxUnavailable(
            "openpyxl is required for .xlsx export. Install it with 'pip install openpyxl'."
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "timeline"

    headers = list(TIMELINE_COLUMNS)
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F3864")
    for column, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    tint = {
        "instruction": "FFF2CC",
        "fixation": "E2EFDA",
        "stimulus": "DDEBF7",
        "blank": "F2F2F2",
    }
    for row in timeline.rows():
        sheet.append([row[key] for key in headers])
        fill = tint.get(str(row["event_type"]))
        if fill:
            for column in range(1, len(headers) + 1):
                sheet.cell(row=sheet.max_row, column=column).fill = PatternFill("solid", fgColor=fill)

    for column, header in enumerate(headers, start=1):
        widest = max([len(header)] + [len(str(r[header])) for r in timeline.rows()] or [10])
        sheet.column_dimensions[get_column_letter(column)].width = min(46, max(10, widest + 2))
    sheet.freeze_panes = "A2"

    summary = timeline.summary()
    sheet.append([])
    sheet.append(["participant", summary["participant"]])
    sheet.append(["total duration (s)", summary["duration_s"]])
    sheet.append(["events", summary["n_events"]])
    sheet.append(["stimuli", summary["n_stimuli"]])

    workbook.save(path)
    return path


def write_settings_json(
    config: BuildConfig,
    path: str | Path,
    *,
    participant: str,
    stimulus_ids: Sequence[str],
    assignment_meta: dict | None = None,
    ffmpeg_version: str = "",
) -> Path:
    """Write the complete, reproducible description of one build."""
    from .. import __version__

    extra = {
        "generator": "stim-concat",
        "version": __version__,
        "built_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "participant": participant,
        "stimulus_ids": [str(s) for s in stimulus_ids],
        "ffmpeg": ffmpeg_version,
        "assignment": assignment_meta or {},
    }
    return config.to_json(path, extra=extra)


def write_summary_csv(results: Iterable[dict], path: str | Path) -> Path:
    """Write the per-participant build summary table."""
    rows = list(results)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "participant",
        "status",
        "video",
        "duration_s",
        "n_events",
        "n_stimuli",
        "stimulus_ids",
        "message",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


def write_summary_markdown(summary: dict, path: str | Path) -> Path:
    """Write a short human-readable build report."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# stim-concat build summary",
        "",
        f"- Generated: {summary.get('built_utc', '')}",
        f"- Stimulus folder: `{summary.get('stimulus_folder', '')}`",
        f"- Assignment sheet: `{summary.get('assignment_csv', '')}`",
        f"- Output folder: `{summary.get('output_folder', '')}`",
        f"- Participants built: {summary.get('n_built', 0)} of {summary.get('n_total', 0)}",
        f"- Total video duration: {summary.get('total_duration_s', 0):.1f} s",
        f"- FFmpeg: {summary.get('ffmpeg', 'unknown')}",
        "",
        "## Participants",
        "",
        "| Participant | Status | Duration (s) | Events | Stimuli |",
        "| --- | --- | ---: | ---: | ---: |",
    ]
    for row in summary.get("participants", []):
        lines.append(
            f"| {row.get('participant', '')} | {row.get('status', '')} | "
            f"{row.get('duration_s', 0)} | {row.get('n_events', 0)} | {row.get('n_stimuli', 0)} |"
        )
    failures = [r for r in summary.get("participants", []) if r.get("status") != "ok"]
    if failures:
        lines += ["", "## Problems", ""]
        for row in failures:
            lines.append(f"- **{row.get('participant')}**: {row.get('message', '')}")
    lines += ["", "## Settings", "", "```json", json.dumps(summary.get("config", {}), indent=2), "```", ""]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
